"""
core/notifier.py  ·  Biótica Consultores
=========================================
Envía notificación por Gmail con los datos del lead.
Adjunta un mini-Excel con la fila del cliente.

Configuración requerida en .env:
    SMTP_USER     = tucorreo@gmail.com          ← correo origen
    SMTP_PASS     = xxxx xxxx xxxx xxxx         ← contraseña de aplicación Gmail
    EMAIL_DESTINO = secretaria@bioticaconsultores.com  ← destino

Para obtener la contraseña de aplicación de Gmail:
  1. Activa la verificación en 2 pasos en tu cuenta Google.
  2. Ve a: Google Account → Seguridad → Contraseñas de aplicaciones.
  3. Crea una para "Correo / Windows" y copia las 16 letras.
  4. Pega esas 16 letras (sin espacios) en SMTP_PASS del .env.
"""

import io
import os
import smtplib
from datetime import datetime
from email.message import EmailMessage

import openpyxl
from dotenv import load_dotenv

load_dotenv()

EMAIL_DESTINO = os.getenv("EMAIL_DESTINO", "secretaria@bioticaconsultores.com")
SMTP_USER     = os.getenv("SMTP_USER", "")
SMTP_PASS     = os.getenv("SMTP_PASS", "")
EMAIL_ORIGEN  = SMTP_USER  # el origen ES el usuario autenticado en Gmail


# ─────────────────────────────────────────────────────────────────────────────
# Función principal
# ─────────────────────────────────────────────────────────────────────────────
def enviar_notificacion_lead(datos_lead: dict) -> bool:
    """
    Envía un correo a EMAIL_DESTINO con:
      - Cuerpo HTML con la tabla del lead
      - Archivo Excel adjunto con Fecha, Nombre, Contacto, Servicio,
        Urgencia, Proyecto, Ubicación, Estado

    Si SMTP_USER / SMTP_PASS no están configurados, simula en consola.
    """
    if not SMTP_USER or not SMTP_PASS:
        _simular(datos_lead)
        return True

    try:
        msg = EmailMessage()
        msg["Subject"] = (
            f"🌱 Nuevo Lead Biótica – {datos_lead.get('clasificacion', 'N/A')} "
            f"[{datos_lead.get('urgencia', 'Normal')}]"
        )
        msg["From"] = EMAIL_ORIGEN
        msg["To"]   = EMAIL_DESTINO

        # ── Cuerpo HTML ───────────────────────────────────────────────
        msg.add_alternative(_html_body(datos_lead), subtype="html")

        # ── Excel adjunto ─────────────────────────────────────────────
        excel_bytes = _generar_excel(datos_lead)
        fecha_str   = datetime.now().strftime("%Y%m%d_%H%M")
        msg.add_attachment(
            excel_bytes,
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"lead_biotica_{fecha_str}.xlsx",
        )

        # ── Envío Gmail (TLS puerto 587) ──────────────────────────────
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.ehlo()
            server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.send_message(msg)

        print(f"[notifier] ✅ Correo enviado a {EMAIL_DESTINO}")
        return True

    except Exception as e:
        print(f"[notifier] ⚠️  Error Gmail: {e}")
        _simular(datos_lead)
        return False


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def _generar_excel(datos: dict) -> bytes:
    """Genera un Excel en memoria con una fila de datos del lead."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lead"

    headers = ["Fecha", "Nombre", "Contacto", "Servicio",
               "Urgencia", "Proyecto", "Ubicación", "Estado"]

    # Estilo del encabezado
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="334737")
    header_font = Font(bold=True, color="FFFFFF")

    ws.append(headers)
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Fila de datos
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        datos.get("nombre",        "N/A"),
        datos.get("contacto",      "N/A"),
        datos.get("clasificacion", "N/A"),
        datos.get("urgencia",      "Normal"),
        datos.get("proyecto",      "N/A"),
        datos.get("ubicacion",     "N/A"),
        datos.get("siguiente_paso","N/A"),
    ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _html_body(d: dict) -> str:
    urgencia_color = "#e65100" if d.get("urgencia") == "Alta" else "#2e7d32"
    return f"""
    <html><body style="font-family:Arial,sans-serif;max-width:600px;margin:auto;">
      <div style="background:#334737;padding:20px;border-radius:8px 8px 0 0;">
        <h2 style="color:#54B435;margin:0;">🌿 Nuevo Lead Calificado</h2>
        <p style="color:rgba(255,255,255,0.7);margin:4px 0 0;">Biótica Consultores</p>
      </div>
      <div style="background:#f4f8f4;padding:24px;border:1px solid #d6e8d0;">
        <table style="width:100%;border-collapse:collapse;">
          {_fila("📅 Fecha",      datetime.now().strftime("%d/%m/%Y %H:%M"))}
          {_fila("👤 Nombre",     d.get("nombre","N/A"))}
          {_fila("📞 Contacto",   d.get("contacto","N/A"))}
          {_fila("📁 Servicio",   d.get("clasificacion","N/A"))}
          {_fila("⚡ Urgencia",   d.get("urgencia","Normal"), color=urgencia_color)}
          {_fila("🏗️ Proyecto",   d.get("proyecto","N/A"))}
          {_fila("📍 Ubicación",  d.get("ubicacion","N/A"))}
          {_fila("✅ Estado",     d.get("siguiente_paso","N/A"))}
        </table>
        <div style="background:white;border:1px solid #d6e8d0;border-radius:6px;padding:14px;margin-top:16px;">
          <strong style="color:#334737;">Resumen técnico:</strong>
          <p style="color:#555;margin:6px 0 0;">{d.get("info_tecnica","Sin información adicional.")}</p>
        </div>
      </div>
      <div style="background:#334737;padding:12px 20px;border-radius:0 0 8px 8px;text-align:center;">
        <p style="color:rgba(255,255,255,0.5);font-size:12px;margin:0;">
          Mensaje automático · Biótica Consultores LTDA · Floridablanca, Santander · Colombia
        </p>
      </div>
    </body></html>
    """


def _fila(label: str, valor: str, color: str = "#333") -> str:
    return f"""
      <tr>
        <td style="padding:8px 12px;border-bottom:1px solid #e0ece0;
                   font-weight:bold;color:#334737;width:40%;font-size:13px;">{label}</td>
        <td style="padding:8px 12px;border-bottom:1px solid #e0ece0;
                   color:{color};font-size:13px;">{valor}</td>
      </tr>"""


def _simular(datos: dict):
    sep = "=" * 55
    print(f"\n{sep}")
    print("📧  SIMULACIÓN CORREO  (configura SMTP_USER y SMTP_PASS en .env)")
    print(f"  Para: {EMAIL_DESTINO}")
    print(f"  Asunto: Nuevo Lead – {datos.get('clasificacion')} [{datos.get('urgencia')}]")
    print(f"  Nombre: {datos.get('nombre')} | Contacto: {datos.get('contacto')}")
    print(f"  Servicio: {datos.get('clasificacion')} | Urgencia: {datos.get('urgencia')}")
    print(f"  Proyecto: {datos.get('proyecto')} | Ubicación: {datos.get('ubicacion')}")
    print(sep + "\n")
